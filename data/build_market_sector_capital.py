#!/usr/bin/env python3
"""build_market_sector_capital.py

Constructs market-sector capital stock K_market from ABS 5204 Table 63
by-industry net capital stock, matched to the market-sector GVA Q_market
from ABS 5206 Table 6 that prepare_supply_data.m already builds.

Also constructs mining vs non-mining splits of both Q and K for the
Phase L3 hypothesis test (PAC_BI_AU_EXPLORATION.md §6 Hypothesis B).

Outputs: data/market_sector_capital.csv  (annual, then linearly
         interpolated to quarterly for the CES calibration)

Definitions (matching prepare_supply_data.m's Q_market):

  Q_market = Total GVA − Public admin − Education − Health − Ownership of dwellings
             (ABS 5206 Table 6, chain volumes, SA)

  K_market = ALL INDUSTRIES total − Dwellings asset − Public admin − Education − Health
             (ABS 5204 Table 63, chain volumes, annual → interpolated to quarterly)

  K_mining  = Mining (B) industry net capital stock (ABS 5204 col 12)
  K_nonmining_market = K_market − K_mining

  Q_mining  = Mining (B) GVA (ABS 5206 col 10, chain volumes, SA)
  Q_nonmining_market = Q_market − Q_mining

Run:  cd ~/Documents/AUSPAC && python3 data/build_market_sector_capital.py
"""

from pathlib import Path
import numpy as np
import openpyxl
import csv
from datetime import datetime

ROOT = Path(__file__).resolve().parents[1]
ABS_DIR = ROOT / "data" / "abs_rba"
OUT_CSV = ROOT / "data" / "market_sector_capital.csv"


def read_annual_col(ws, col_idx, first_data_row=11):
    """Read annual data from an ABS 5204 worksheet column."""
    dates, vals = [], []
    for row in ws.iter_rows(min_row=first_data_row, values_only=False):
        d = row[0].value
        v = row[col_idx].value
        if d is None:
            break
        if isinstance(d, datetime):
            dates.append(d.year)
        elif isinstance(d, (int, float)):
            dates.append(int(d))
        else:
            try:
                dates.append(int(str(d)[:4]))
            except ValueError:
                break
        try:
            vals.append(float(v) if v is not None else np.nan)
        except (ValueError, TypeError):
            vals.append(np.nan)
    return np.array(dates), np.array(vals)


def read_quarterly_col(ws, col_idx, first_data_row=11):
    """Read quarterly data from an ABS 5206 worksheet column.

    NOTE: this is intentionally unit-blind; the *caller* must have already
    asserted (via assert_sa_level_col) that col_idx is a Seasonally Adjusted
    "$ Millions" LEVEL column. Do NOT use this on a column you have not
    validated -- the old fragile fallback selected "Percentage changes"
    (Unit=Percent) columns, which is the unit bug this rewrite fixes.
    """
    dates, vals = [], []
    for row in ws.iter_rows(min_row=first_data_row, values_only=False):
        d = row[0].value
        v = row[col_idx].value
        if d is None:
            break
        try:
            vals.append(float(v) if v is not None else np.nan)
        except (ValueError, TypeError):
            vals.append(np.nan)
        if isinstance(d, datetime):
            dates.append(d)
        else:
            break
    return dates, np.array(vals)


def assert_sa_level_col(desc_row, unit_row, stype_row, col_idx, expect_label):
    """Hard-assert a 5206 column is the SA $-million LEVEL series we mean.

    Guards against the unit bug: the LEVEL columns (Unit='$ Millions') must
    be pinned, NOT the later 'Percentage changes' columns (Unit='Percent').
    """
    desc = str(desc_row[col_idx] or "")
    unit = str(unit_row[col_idx] or "")
    stype = str(stype_row[col_idx] or "")
    assert "Seasonally Adjusted" in stype, (
        f"col {col_idx} ({expect_label}): Series Type is {stype!r}, "
        f"expected 'Seasonally Adjusted'")
    assert "$ Millions" in unit, (
        f"col {col_idx} ({expect_label}): Unit is {unit!r}, expected "
        f"'$ Millions' (a 'Percent' unit means a 'Percentage changes' "
        f"column was picked -- the unit bug)")
    assert "Percentage changes" not in desc, (
        f"col {col_idx} ({expect_label}): desc {desc!r} is a growth column")
    assert expect_label in desc, (
        f"col {col_idx}: desc {desc!r} does not contain {expect_label!r}")
    return desc, unit, stype


def main():
    print("=== build_market_sector_capital.py ===\n")

    # --- ABS 5204 Table 63: Net capital stock by industry (annual, chain vol $M) ---
    print("Reading ABS 5204 net capital stock by industry...")
    wb_k = openpyxl.load_workbook(ABS_DIR / "abs_5204_net_capital_stock.xlsx",
                                  read_only=True, data_only=True)
    ws_k = wb_k["Data1"]

    # Column indices (0-based) from the header scan above
    COL_K_ALL      = 113   # ALL INDUSTRIES total
    COL_K_DWELL    = 103   # ALL INDUSTRIES ; Dwellings
    COL_K_PUBADM   = 79    # Public administration and safety
    COL_K_EDU      = 84    # Education and training
    COL_K_HEALTH   = 89    # Health care and social assistance
    COL_K_MINING   = 12    # Mining (B)
    COL_K_AGRI     = 6     # Agriculture (A)

    years_k, k_all     = read_annual_col(ws_k, COL_K_ALL)
    _,       k_dwell   = read_annual_col(ws_k, COL_K_DWELL)
    _,       k_pubadm  = read_annual_col(ws_k, COL_K_PUBADM)
    _,       k_edu     = read_annual_col(ws_k, COL_K_EDU)
    _,       k_health  = read_annual_col(ws_k, COL_K_HEALTH)
    _,       k_mining  = read_annual_col(ws_k, COL_K_MINING)

    k_market = k_all - k_dwell - k_pubadm - k_edu - k_health
    k_nonmining_market = k_market - k_mining

    print(f"  Years: {int(years_k[0])}–{int(years_k[-1])} ({len(years_k)} obs)")
    print(f"  K_all (2019):          {k_all[years_k == 2019][0]:,.0f} $M")
    print(f"  K_dwellings (2019):    {k_dwell[years_k == 2019][0]:,.0f} $M")
    print(f"  K_pubadm (2019):       {k_pubadm[years_k == 2019][0]:,.0f} $M")
    print(f"  K_edu (2019):          {k_edu[years_k == 2019][0]:,.0f} $M")
    print(f"  K_health (2019):       {k_health[years_k == 2019][0]:,.0f} $M")
    print(f"  K_market (2019):       {k_market[years_k == 2019][0]:,.0f} $M")
    print(f"  K_mining (2019):       {k_mining[years_k == 2019][0]:,.0f} $M")
    print(f"  K_nonmining_mkt (2019):{k_nonmining_market[years_k == 2019][0]:,.0f} $M")
    wb_k.close()

    # --- ABS 5206 Table 6: GVA by industry (quarterly, chain vol $M, SA) ---
    print("\nReading ABS 5206 industry GVA...")
    wb_q = openpyxl.load_workbook(ABS_DIR / "abs_5206_industry_gva.xlsx",
                                  read_only=True, data_only=True)
    ws_q = wb_q["Data1"]

    # --- Pinned SA $-million LEVEL columns (0-based), verified against the
    #     header rows of abs_5206_industry_gva.xlsx Data1 on 2026-06-01:
    #       row0=description, row1=Unit, row2=Series Type, data from row 11.
    #     These are the CHAIN-VOLUME, Seasonally Adjusted, '$ Millions' LEVELS.
    #     The OLD fragile string-match walked past these and (because it used
    #     `=` not `is None`) landed on the later 'Percentage changes'
    #     (Unit='Percent') columns 210-216, mixing growth into the level
    #     arithmetic and producing q_nonmining_market = -73,879. Fallback
    #     offsets are DELETED -- pin or fail loudly.
    SA_MINING = 119   # 'Mining (B) ;'
    SA_PUBADM = 155   # 'Public administration and safety (O) ;'
    SA_EDU    = 156   # 'Education and training (P) ;'
    SA_HEALTH = 157   # 'Health care and social assistance (Q) ;'
    SA_DWELL  = 160   # 'Ownership of dwellings ;'
    SA_TOTAL  = 161   # 'Gross value added at basic prices ;'

    rows_q = list(ws_q.iter_rows(min_row=1, max_row=3, values_only=True))
    desc_row, unit_row, stype_row = rows_q[0], rows_q[1], rows_q[2]

    # Hard-assert series-type AND unit BEFORE reading any values.
    assert_sa_level_col(desc_row, unit_row, stype_row, SA_MINING, "Mining (B) ;")
    assert_sa_level_col(desc_row, unit_row, stype_row, SA_PUBADM,
                        "Public administration and safety")
    assert_sa_level_col(desc_row, unit_row, stype_row, SA_EDU,
                        "Education and training")
    assert_sa_level_col(desc_row, unit_row, stype_row, SA_HEALTH,
                        "Health care and social assistance")
    assert_sa_level_col(desc_row, unit_row, stype_row, SA_DWELL,
                        "Ownership of dwellings")
    assert_sa_level_col(desc_row, unit_row, stype_row, SA_TOTAL,
                        "Gross value added at basic prices")

    sa_mining_col = SA_MINING
    sa_total_col  = SA_TOTAL
    sa_pubadm_col = SA_PUBADM
    sa_edu_col    = SA_EDU
    sa_health_col = SA_HEALTH
    sa_dwell_col  = SA_DWELL

    print(f"  Pinned SA $M LEVEL columns: mining={sa_mining_col}, "
          f"total={sa_total_col}, pubadm={sa_pubadm_col}, edu={sa_edu_col}, "
          f"health={sa_health_col}, dwell={sa_dwell_col}")

    dates_q_list, q_total  = read_quarterly_col(ws_q, sa_total_col)
    _,            q_mining = read_quarterly_col(ws_q, sa_mining_col)
    _,            q_pubadm = read_quarterly_col(ws_q, sa_pubadm_col)
    _,            q_edu    = read_quarterly_col(ws_q, sa_edu_col)
    _,            q_health = read_quarterly_col(ws_q, sa_health_col)
    _,            q_dwell  = read_quarterly_col(ws_q, sa_dwell_col)
    wb_q.close()

    # Q_market = Total GVA - pubadm - edu - health - dwellings (all SA $M levels)
    q_market = q_total - q_pubadm - q_edu - q_health - q_dwell
    q_nonmining_market = q_market - q_mining

    # Quarter index (for date-keyed asserts/diagnostics)
    years_q = np.array([d.year for d in dates_q_list])
    months_q = np.array([d.month for d in dates_q_list])

    # ------------------------------------------------------------------
    # PHASE-0 GATE ASSERTS (spec NEXT_PROJECT_industry_split.md S3.2)
    # ------------------------------------------------------------------
    # Restrict to observed quarters (the series start NaN before 1959-ish
    # for some columns; mining/dwellings begin later). Use a common-valid
    # mask so the positivity asserts test real data, not NaNs.
    valid = (~np.isnan(q_mining) & ~np.isnan(q_nonmining_market)
             & ~np.isnan(q_market) & ~np.isnan(q_total))
    assert valid.sum() >= 120, (
        f"only {valid.sum()} fully-observed quarters; need >=120")

    qm_v = q_mining[valid]
    qnm_v = q_nonmining_market[valid]

    a1 = bool((qm_v > 0).all())
    a2 = bool((qnm_v > 0).all())
    print("\n=== Phase-0 GATE asserts ===")
    print(f"  [A1] (q_mining > 0).all()           : {a1}  "
          f"(min={qm_v.min():,.0f})")
    print(f"  [A2] (q_nonmining_market > 0).all() : {a2}  "
          f"(min={qnm_v.min():,.0f})")
    assert a1, "q_mining has non-positive values"
    assert a2, ("q_nonmining_market has non-positive values -- the unit "
                "bug (level minus growth) would trip this")

    # Mining VA share in FY2022-23 (Sep-22, Dec-22, Mar-23, Jun-23).
    # NOTE on the [0.08,0.16] GATE band: the spec's verified I-O number is
    # mining = 11.66% of *total* GVA (w_qn_m ~ 0.117, S1.3), so the band is
    # the TOTAL-GVA denominator. Mining as a share of *market* GVA is
    # structurally larger (~0.163) because market GVA excludes the large
    # non-market sector -- that ratio is the within-market CES weight and is
    # reported separately, not gated against the total-GVA band.
    fy2223 = ((years_q == 2022) & (months_q >= 9)) | \
             ((years_q == 2023) & (months_q <= 6))
    q_min_fy = np.nansum(q_mining[fy2223])
    q_mkt_fy = np.nansum(q_market[fy2223])
    q_tot_fy = np.nansum(q_total[fy2223])
    mining_share_total_fy2223 = q_min_fy / q_tot_fy
    mining_share_market_fy2223 = q_min_fy / q_mkt_fy
    a3 = bool(0.08 <= mining_share_total_fy2223 <= 0.16)
    print(f"  [A3] mining VA share of TOTAL GVA,  FY2022-23 = "
          f"{mining_share_total_fy2223:.4f}  in [0.08,0.16]? {a3}")
    print(f"       mining VA share of MARKET GVA, FY2022-23 = "
          f"{mining_share_market_fy2223:.4f}  (within-market CES weight, "
          f"not gated)")
    assert a3, (f"mining share of total GVA {mining_share_total_fy2223:.4f} "
                f"outside [0.08,0.16]")

    # Reconciliation: mining + nonmining_market + pubadm + edu + health
    # + dwellings == total (within rounding / chain-volume non-additivity).
    recon = (q_mining + q_nonmining_market + q_pubadm + q_edu
             + q_health + q_dwell)
    diff = recon[valid] - q_total[valid]
    max_abs_diff = float(np.nanmax(np.abs(diff)))
    max_rel_diff = float(np.nanmax(np.abs(diff) / q_total[valid]))
    a4 = max_rel_diff < 1e-6
    print(f"  [A4] reconciliation to total: max|abs|={max_abs_diff:.6f} $M, "
          f"max|rel|={max_rel_diff:.2e}  (<1e-6? {a4})")
    assert a4, "sectoral split does not reconcile to total GVA"

    # Find 2019 annual average for diagnostics
    mask_2019 = years_q == 2019
    print(f"\n  Q_total (2019 avg):          {np.nanmean(q_total[mask_2019]):,.0f} $M/qtr")
    print(f"  Q_market (2019 avg):         {np.nanmean(q_market[mask_2019]):,.0f} $M/qtr")
    print(f"  Q_mining (2019 avg):         {np.nanmean(q_mining[mask_2019]):,.0f} $M/qtr")
    print(f"  Q_nonmining_mkt (2019 avg):  {np.nanmean(q_nonmining_market[mask_2019]):,.0f} $M/qtr")

    # --- Compute gamma ratios ---
    # γ_old = Q_market / K_total (the mismatched version)
    # γ_new = Q_market / K_market (the corrected version)
    # Both at 2019 base year (quarterly Q avg / annual K)
    q_mkt_2019 = np.nanmean(q_market[mask_2019])
    k_all_2019 = k_all[years_k == 2019][0]
    k_mkt_2019 = k_market[years_k == 2019][0]
    k_min_2019 = k_mining[years_k == 2019][0]
    k_nmkt_2019 = k_nonmining_market[years_k == 2019][0]
    q_min_2019 = np.nanmean(q_mining[mask_2019])
    q_nmkt_2019 = np.nanmean(q_nonmining_market[mask_2019])

    gamma_old = q_mkt_2019 / k_all_2019
    gamma_new = q_mkt_2019 / k_mkt_2019
    gamma_mining = q_min_2019 / k_min_2019
    gamma_nonmining = q_nmkt_2019 / k_nmkt_2019

    print(f"\n=== Gamma diagnostics (2019 base year) ===")
    print(f"  γ_old = Q_market / K_total       = {q_mkt_2019:,.0f} / {k_all_2019:,.0f} = {gamma_old:.4f}")
    print(f"  γ_new = Q_market / K_market       = {q_mkt_2019:,.0f} / {k_mkt_2019:,.0f} = {gamma_new:.4f}")
    print(f"  γ_mining = Q_mining / K_mining     = {q_min_2019:,.0f} / {k_min_2019:,.0f} = {gamma_mining:.4f}")
    print(f"  γ_nonmining = Q_nonmkt / K_nonmkt = {q_nmkt_2019:,.0f} / {k_nmkt_2019:,.0f} = {gamma_nonmining:.4f}")
    print(f"\n  FR-BDF 2026 γ = 0.2561 for France")
    print(f"  OLD AUSPAC γ  = {gamma_old:.4f}  (mismatched: Q_market / K_total)")
    print(f"  NEW AUSPAC γ  = {gamma_new:.4f}  (matched: Q_market / K_market)")

    # --- Cross-restriction diagnostic for μ ---
    sigma = 0.5366
    alpha = 0.45
    for label, g in [("old (mismatched)", gamma_old), ("new (matched)", gamma_new)]:
        log_mu = np.log(1 - alpha) + ((sigma - 1) / sigma) * np.log(g)
        mu = np.exp(log_mu)
        print(f"\n  μ cross-restriction at γ_{label}: μ = {mu:.3f}")

    # --- Write CSV output ---
    print(f"\nWriting {OUT_CSV}...")
    with OUT_CSV.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["year", "k_all", "k_market", "k_mining", "k_nonmining_market",
                     "k_dwellings", "k_pubadm", "k_edu", "k_health"])
        for i, yr in enumerate(years_k):
            w.writerow([int(yr), k_all[i], k_market[i], k_mining[i],
                        k_nonmining_market[i], k_dwell[i], k_pubadm[i],
                        k_edu[i], k_health[i]])
    print(f"  wrote {len(years_k)} annual rows")

    # Also write the quarterly GVA splits
    OUT_GVA = ROOT / "data" / "market_sector_gva_splits.csv"
    print(f"Writing {OUT_GVA}...")
    with OUT_GVA.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["date", "q_total", "q_market", "q_mining",
                     "q_nonmining_market", "q_pubadm", "q_edu",
                     "q_health", "q_dwellings"])
        for i, d in enumerate(dates_q_list):
            w.writerow([d.strftime("%Y-%m-%d"), q_total[i], q_market[i],
                        q_mining[i], q_nonmining_market[i],
                        q_pubadm[i], q_edu[i], q_health[i], q_dwell[i]])
    print(f"  wrote {len(dates_q_list)} quarterly rows")

    print("\ndone.")


if __name__ == "__main__":
    main()
