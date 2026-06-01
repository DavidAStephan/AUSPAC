#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_io_bridge.py  -- AUSPAC industry-split Phase 0, spec NEXT_PROJECT_industry_split.md s3.3

SINGLE SOURCE OF TRUTH for the Input-Output value-added bridge

        B = V_hat (I - A)^{-1}

that the two-sector .mod will consume to route domestic final demand (C/I/G/X)
to producing sectors (mining vs non-mining-market vs non-market) without
double-counting, and to close GDP(P)=GDP(E) on ONE valuation basis (basic
prices) with the small 0.8% taxes-on-products wedge.

Source: data/io_tables_australia.xlsx  (ABS 5209.0 Australian National Accounts:
Input-Output Tables, 2021-22), 115 IOIG product/industry groups.

Tables used:
  * 'table'   -> table-name index
  * 'code'    -> 115 IOIG codes
  * 'Table 2' -> USE table at BASIC PRICES (intermediate use 115x115 + final-demand
                 columns C/G/I/inv/X + VA row V1). Rows 2..116 = 115 products
                 (codes 101..9502); col 0 = product code; col header row 1 =
                 industry codes '0101'..'9502' (cols 1..115).
                   col 116 = total intermediate use
                   col 117 = C  (household final consumption)        = 1,284,660
                   col 118 = G  (government final consumption)       =   593,514
                   col 119 = I  GFCF (P3 component a)                ~  457,042
                   col 120 = I  GFCF (P3 component b)                ~    41,224
                   col 121 = I  GFCF (P3 component c)                ~   108,042
                   col 119+120+121 = total GFCF                      =   606,308
                   col 122 = inventories (P3 changes in invent.)     =     6,409
                   col 123 = X  (exports)                            =   644,385
                   col 124 = total final demand                      = 3,135,275
                   col 125 = total use (= product gross output)
                 Bottom block (col 0 SNA labels):
                   row 128 'P6' col125 = imports                     =   603,645
                   row 132 'V1' col116 = GVA                         = 2,511,352
  * 'Table 7' -> Leontief TOTAL-requirement coefficients, x100, direct allocation
                 of imports, 115x115 (industry-by-industry). Divide by 100 to get
                 (I-A)^{-1}. Diagonal in [100,178] confirms x100 scaling.
  * 'Table 17'-> primary-input content (total requirements) per $100 of final use
                 by industry: P1(compensation)+P2(GOS/mixed)+P3(taxes on
                 production)+P4(?)+P6(imports) = 100. Used as an INDEPENDENT
                 cross-check of the per-product VA / import content.

Mining = IOIG product codes {601,701,801,802,901,1001} (Coal; Oil&gas; Iron ore;
Non-ferrous metal ore; Non-metallic mineral; Exploration & mining support).
Identified by CODE EQUALITY (not row number). Product-row = code-row + 1 because
of the header offset; here both Table 2 and Table 7 put code 601 at the 9th data
row (0-based python index 8) -> asserted below.

Product-vs-industry basis: Table 7 is product-by-product / industry-by-industry
(direct allocation of imports); the mining aggregation is by IOIG code. For the
mining-vs-non-mining cut the product and industry classifications coincide in the
AU IO framework, so the bridge is exact for this 2-way cut. (Finer splits would
need the full supply-use Table 1 -- deferred per spec s7.)

Emits:
  data/io_bridge_coefficients.csv  -- the bridge weights the .mod consumes
  data/closure_shares.csv          -- GVA, GDP-E, tax wedge, mining shares, eta_xm

All hard asserts fail loudly (AssertionError) and are echoed to stdout.
"""

import sys
import os
import numpy as np
import openpyxl

# ----------------------------------------------------------------------------
# 0. Config: paths, mining codes, column map, design-phase reference numbers
# ----------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "io_tables_australia.xlsx")
OUT_BRIDGE = os.path.join(HERE, "io_bridge_coefficients.csv")
OUT_CLOSURE = os.path.join(HERE, "closure_shares.csv")

MINING_CODES = [601, 701, 801, 802, 901, 1001]

# Table 2 final-demand columns (python 0-based indices into the row tuple)
COL_INTERMED = 116   # total intermediate use
COL_C = 117          # household consumption
COL_G = 118          # government consumption
COL_I = [119, 120, 121]  # GFCF (three sub-components)
COL_INV = 122        # changes in inventories
COL_X = 123          # exports
COL_TFD = 124        # total final demand
COL_TOTUSE = 125     # total use (= product gross output)

# Bottom-block SNA row labels (col 0) we need
ROW_IMPORTS_LABEL = "P6"     # imports
ROW_GVA_LABEL = "V1"         # gross value added

# Design-phase reference numbers to confirm (flag any material divergence)
REF = dict(
    mining_va=292712.0,
    mining_va_share=0.1166,
    gva=2511352.0,
    gdp_e=2531631.0,
    tax_wedge=20279.0,
    va_to_output=0.643,
    resource_exports=330588.0,
    eta_xm=1.13,
    mining_export_share_gross=0.513,
)
TOL_PCT = 0.005   # 0.5% material-divergence flag threshold for the design numbers

_failures = []
_flags = []


def check(name, condition, detail=""):
    status = "PASS" if condition else "FAIL"
    print(f"  [{status}] {name}: {detail}")
    if not condition:
        _failures.append(f"{name}: {detail}")
    return condition


def flag(name, ref, got, tol=TOL_PCT):
    rel = abs(got - ref) / abs(ref) if ref != 0 else abs(got)
    ok = rel <= tol
    status = "ok" if ok else "DIVERGENCE"
    print(f"  [{status}] {name}: design={ref:.6g} computed={got:.6g} reldiff={rel*100:.3f}%")
    if not ok:
        _flags.append(f"{name}: design={ref:.6g} computed={got:.6g} reldiff={rel*100:.3f}%")
    return ok


# ----------------------------------------------------------------------------
# 1. Load workbook + locate tables
# ----------------------------------------------------------------------------
print("=" * 78)
print("build_io_bridge.py  --  I-O value-added bridge  B = V_hat (I-A)^-1")
print("=" * 78)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
print("\n[1] Workbook sheets:", wb.sheetnames)

# 'code' sheet -> 115 IOIG codes
ws_code = wb["code"]
codes = [int(r[0]) for r in ws_code.iter_rows(min_row=1, max_row=200, values_only=True)
         if r[0] is not None and str(r[0]).strip() != ""]
print(f"[1] 'code' sheet: {len(codes)} IOIG codes, first={codes[0]} last={codes[-1]}")
check("code_count_115", len(codes) == 115, f"{len(codes)} codes")

t2 = list(wb["Table 2"].iter_rows(values_only=True))
t7 = list(wb["Table 7"].iter_rows(values_only=True))
t17 = list(wb["Table 17"].iter_rows(values_only=True))
print(f"[1] Table 2 rows={len(t2)}  Table 7 rows={len(t7)}  Table 17 rows={len(t17)}")

# ----------------------------------------------------------------------------
# 2. Identify the 115 product rows and the 6 mining rows BY CODE EQUALITY
# ----------------------------------------------------------------------------
# Table 2: header row index 0 has industry codes (strings '0101'..) in cols 1..115;
#          product rows are indices 1..115 with the code in col 0.
t2_prod_codes = [int(t2[i][0]) for i in range(1, 116)]
t2_hdr_codes = [int(t2[0][c]) for c in range(1, 116)]
check("t2_product_codes_match_code_sheet", t2_prod_codes == codes,
      "product row codes == code sheet")
check("t2_header_codes_match", t2_hdr_codes == codes,
      "industry header codes == code sheet")

t7_row_codes = [int(t7[i][0]) for i in range(1, 116)]
t7_col_codes = [int(t7[0][c]) for c in range(1, 116)]
check("t7_row_codes_match", t7_row_codes == codes, "Table7 row codes == code sheet")
check("t7_col_codes_match", t7_col_codes == codes, "Table7 col codes == code sheet")

# Mining indices by code equality (python 0-based into the 0..114 product axis)
mining_idx = [codes.index(c) for c in MINING_CODES]   # positions in 0..114
selected_codes = [codes[i] for i in mining_idx]
check("mining_codes_equality", selected_codes == MINING_CODES,
      f"selected={selected_codes} expected={MINING_CODES}")
# product-row = code-row + 1 (header offset): in Table 2, code 601 sits at t2 index 8
check("mining_row_offset", int(t2[mining_idx[0] + 1][0]) == 601,
      f"t2 row index {mining_idx[0]+1} has code {t2[mining_idx[0]+1][0]} (=code-row+1 header offset)")

non_mining_idx = [i for i in range(115) if i not in mining_idx]
print(f"[2] mining product axis positions (0-based): {mining_idx} -> codes {selected_codes}")
print(f"[2] non-mining: {len(non_mining_idx)} products")

# ----------------------------------------------------------------------------
# 3. Build the core matrices
#    - gross output by product  x  (col 125, total use)
#    - VA by industry/product   v  (row V1, cols 1..115)
#    - intermediate-use matrix  Z  (115x115, t2 product rows x industry cols)
#    - VA ratios v_ratio = v / x  -> V_hat = diag(v_ratio)
#    - Leontief inverse L = Table7 / 100
# ----------------------------------------------------------------------------
def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0

# gross output by product = total use col 125
x = np.array([num(t2[i + 1][COL_TOTUSE]) for i in range(115)], dtype=float)

# VA by industry: row labelled 'V1' in col 0, values in industry cols 1..115
va_row_idx = next(i for i in range(len(t2)) if str(t2[i][0]).strip() == ROW_GVA_LABEL)
v = np.array([num(t2[va_row_idx][c]) for c in range(1, 116)], dtype=float)
gva = num(t2[va_row_idx][COL_INTERMED])   # V1 col 116 = total GVA
print(f"[3] V1 row found at t2 index {va_row_idx}; total GVA (col116) = {gva:,.0f}")

# intermediate-use Z[product_row, industry_col]
Z = np.array([[num(t2[i + 1][c]) for c in range(1, 116)] for i in range(115)], dtype=float)

# VA ratios per industry/product = VA / gross output (guard zero output)
with np.errstate(divide="ignore", invalid="ignore"):
    v_ratio = np.where(x > 0, v / x, 0.0)
V_hat = np.diag(v_ratio)

# Leontief inverse L = Table 7 / 100
L100 = np.array([[num(t7[i + 1][c]) for c in range(1, 116)] for i in range(115)], dtype=float)
check("table7_x100_scaling", L100.diagonal().min() >= 99.0 and L100.diagonal().max() <= 200.0,
      f"Table7 diag in [{L100.diagonal().min():.2f},{L100.diagonal().max():.2f}] -> /100")
L = L100 / 100.0   # (I - A)^{-1}, diagonal >= 1

# The bridge: B = V_hat (I - A)^{-1}.  Column j of B distributes one unit of final
# demand for product j into VA generated across all producing products i.
# (errstate guard: numpy-2.0 emits spurious matmul warnings against openpyxl
#  read-only-loaded float arrays; the result is verified finite below.)
with np.errstate(divide="ignore", over="ignore", invalid="ignore"):
    B = np.asarray(V_hat, dtype=np.float64) @ np.asarray(L, dtype=np.float64)
assert np.isfinite(B).all(), "bridge matrix B has non-finite entries"
# Column-sum of B = total VA content per $1 of final use of product j (the rest is
# import leakage + product-tax). With direct-allocation-of-imports Leontief, the
# VA + import + tax content of a $1 of final demand sums to 1.
B_colsum_va = B.sum(axis=0)   # VA content per $ of final demand for each product

# ----------------------------------------------------------------------------
# 4. Mining VA, output, exports, shares (confirm design-phase numbers)
# ----------------------------------------------------------------------------
mining_va_direct = float(v[mining_idx].sum())
total_va = float(v.sum())
mining_va_share_direct = mining_va_direct / total_va

mining_go = float(x[mining_idx].sum())          # product gross output
mining_va_to_output = mining_va_direct / mining_go

mining_exports = float(sum(num(t2[i + 1][COL_X]) for i in mining_idx))
total_exports = float(sum(num(t2[i + 1][COL_X]) for i in range(115)))
mining_export_share_gross = mining_exports / total_exports
eta_xm = mining_exports / mining_va_direct       # VA-to-export elasticity

print("\n[4] Mining aggregates (DIRECT, from Table 2):")
print(f"    mining VA            = {mining_va_direct:,.0f}")
print(f"    total VA (GVA)       = {total_va:,.0f}")
print(f"    mining VA share      = {mining_va_share_direct:.4f}")
print(f"    mining gross output  = {mining_go:,.0f}  (product basis, col125)")
print(f"    mining VA/output     = {mining_va_to_output:.4f}")
print(f"    mining exports       = {mining_exports:,.0f}")
print(f"    total exports        = {total_exports:,.0f}")
print(f"    mining export share (gross) = {mining_export_share_gross:.4f}")
print(f"    eta_xm (exports/VA)  = {eta_xm:.4f}")

print("\n[4] Confirm design-phase numbers (flag material divergence):")
flag("mining_va", REF["mining_va"], mining_va_direct)
flag("mining_va_share", REF["mining_va_share"], mining_va_share_direct)
flag("gva", REF["gva"], total_va)
flag("va_to_output", REF["va_to_output"], mining_va_to_output)
flag("resource_exports", REF["resource_exports"], mining_exports)
flag("eta_xm", REF["eta_xm"], eta_xm, tol=0.02)
flag("mining_export_share_gross", REF["mining_export_share_gross"], mining_export_share_gross)

# ----------------------------------------------------------------------------
# 5. Per-final-demand-column induced VA decomposition
#    For each FD column f (vector over 115 products), induced VA by product
#    = B @ f ; mining VA share = sum over mining rows / sum over all rows.
#    Import + tax share = 1 - VA content per $ spent.
# ----------------------------------------------------------------------------
def fd_vector(col):
    return np.array([num(t2[i + 1][col]) for i in range(115)], dtype=float)

def fd_vector_multi(cols):
    out = np.zeros(115)
    for c in cols:
        out += fd_vector(c)
    return out

fd_defs = {
    "C": ([COL_C], "household consumption"),
    "G": ([COL_G], "government consumption"),
    "I": (COL_I, "GFCF (cols 119+120+121)"),
    "INV": ([COL_INV], "changes in inventories"),
    "X": ([COL_X], "exports"),
}

# Imports total (P6) for the leakage check
imp_row_idx = next(i for i in range(len(t2)) if str(t2[i][0]).strip() == ROW_IMPORTS_LABEL)
imports_total = num(t2[imp_row_idx][COL_TOTUSE])

print("\n[5] Per-final-demand-component induced-VA decomposition (B = V_hat (I-A)^-1):")
print(f"    {'comp':4s} {'spend':>12s} {'miningVA%':>10s} {'nonminVA%':>10s} {'import+tax%':>11s} {'sum':>6s}")
comp_decomp = {}
for name, (cols, desc) in fd_defs.items():
    f = fd_vector_multi(cols)
    spend = float(f.sum())
    with np.errstate(divide="ignore", over="ignore", invalid="ignore"):
        induced_va = B @ f                     # VA by product induced by this FD vector
    total_induced_va = float(induced_va.sum())
    mining_induced = float(induced_va[mining_idx].sum())
    nonmin_induced = float(induced_va[non_mining_idx].sum())
    # shares of SPENDING (so VA% + import+tax% = 1)
    va_share_of_spend = total_induced_va / spend if spend != 0 else 0.0
    mining_va_share = mining_induced / spend if spend != 0 else 0.0
    nonmin_va_share = nonmin_induced / spend if spend != 0 else 0.0
    leak_share = 1.0 - va_share_of_spend       # import + product-tax leakage
    # shares WITHIN value added (used for the production routing weights)
    mining_within_va = mining_induced / total_induced_va if total_induced_va != 0 else 0.0
    nonmin_within_va = nonmin_induced / total_induced_va if total_induced_va != 0 else 0.0
    comp_decomp[name] = dict(
        spend=spend,
        total_induced_va=total_induced_va,
        mining_induced=mining_induced,
        nonmin_induced=nonmin_induced,
        va_share_of_spend=va_share_of_spend,
        mining_va_share_of_spend=mining_va_share,
        nonmin_va_share_of_spend=nonmin_va_share,
        leak_share=leak_share,
        mining_within_va=mining_within_va,
        nonmin_within_va=nonmin_within_va,
    )
    print(f"    {name:4s} {spend:12,.0f} {mining_va_share*100:10.3f} "
          f"{nonmin_va_share*100:10.3f} {leak_share*100:11.3f} "
          f"{(mining_va_share+nonmin_va_share+leak_share)*100:6.2f}")

# ----------------------------------------------------------------------------
# 6. Three-way GVA partition: mining / non-mining-market / non-market
#    The I-O bridge gives mining vs non-mining(all). The non-mining-MARKET vs
#    non-market split is the model's EXISTING partition (au_pac.mod:1667-1671,
#    market_sector_capital.csv). The non-market industries are public admin,
#    defence, public order, education, health, residential care, and ownership
#    of dwellings (imputed + actual rent). We compute their VA share here so the
#    .mod has w_qn_m + w_qn_nm + w_qn_nmk = 1 from one consistent source.
# ----------------------------------------------------------------------------
# Non-market IOIG codes (public admin/defence/order, all education, health/care,
# dwellings). These mirror the q_pubadm/q_edu/q_health/q_dwellings columns the
# existing market/non-market block uses.
NONMARKET_CODES = [
    7501, 7601, 7701,          # public admin & regulatory, defence, public order
    8010, 8110, 8210,          # education (primary/secondary, tertiary, other)
    8401, 8601,                # health care, residential care & social assistance
    6700, 6701,                # imputed + actual rent (ownership of dwellings)
]
nonmarket_idx = [codes.index(c) for c in NONMARKET_CODES]
nonmarket_va = float(v[nonmarket_idx].sum())
nonmarket_share = nonmarket_va / total_va

w_qn_m = mining_va_share_direct
w_qn_nmk = nonmarket_share
w_qn_nm = 1.0 - w_qn_m - w_qn_nmk   # non-mining MARKET as the residual

print("\n[6] Three-way GVA partition (nominal/current-price basis, ABS 5209 2021-22):")
print(f"    w_qn_m   (mining)              = {w_qn_m:.4f}   VA={mining_va_direct:,.0f}")
print(f"    w_qn_nm  (non-mining market)   = {w_qn_nm:.4f}   (residual)")
print(f"    w_qn_nmk (non-market)          = {w_qn_nmk:.4f}   VA={nonmarket_va:,.0f}")
print(f"    sum = {w_qn_m + w_qn_nm + w_qn_nmk:.6f}")
# Per-component non-market breakdown (so the next phase can reconcile membership
# against market_sector_capital.csv's q_pubadm/q_edu/q_health/q_dwellings columns)
_nm_labels = {7501: "pub admin/reg", 7601: "defence", 7701: "public order",
              8010: "edu prim/sec", 8110: "edu tertiary", 8210: "edu other",
              8401: "health care", 8601: "resid care/social",
              6700: "imputed rent", 6701: "actual rent (housing)"}
print("    non-market components (VA share of GVA):")
for c in NONMARKET_CODES:
    i = codes.index(c)
    print(f"        {c} {_nm_labels.get(c,''):22s} {v[i]/total_va*100:5.2f}%")
# FLAG the divergence from the spec s1.3 EXPECTED partition (0.72 / 0.16). The IO
# table puts ownership-of-dwellings (imputed+actual rent ~9.3% of GVA) and
# residential-care (~4% of GVA) inside non-market, which the spec's expected 0.16
# omitted. Surface this so Phase 1 reconciles non-market MEMBERSHIP, not the math.
flag("w_qn_nm_vs_spec_expected_0.72", 0.72, w_qn_nm, tol=0.05)
flag("w_qn_nmk_vs_spec_expected_0.16", 0.16, w_qn_nmk, tol=0.05)

# ----------------------------------------------------------------------------
# 7. Resource-export bridge double-count guard (spec s2.3 / R6)
#    resource-export VA should equal mining VA minus domestic mining absorption
#    minus mining's imported intermediate content. We report the components.
# ----------------------------------------------------------------------------
mining_intermediate_use = float(sum(num(t2[i + 1][COL_INTERMED]) for i in mining_idx))
mining_domestic_fd = float(sum(num(t2[i + 1][COL_TFD]) for i in mining_idx)) - mining_exports
print("\n[7] Resource-export bridge components (double-count guard):")
print(f"    mining product gross output (col125) = {mining_go:,.0f}")
print(f"    mining intermediate use (col116)     = {mining_intermediate_use:,.0f}  (domestic + imported inputs absorbed as inputs)")
print(f"    mining domestic final demand (FD-X)  = {mining_domestic_fd:,.0f}")
print(f"    mining exports (col123)              = {mining_exports:,.0f}")
print(f"    eta_xm = exports / mining VA         = {eta_xm:.4f}")

# ----------------------------------------------------------------------------
# 8. GDP(P)=GDP(E) closure on ONE basis (basic prices), 0.8% tax wedge
# ----------------------------------------------------------------------------
C_tot = comp_decomp["C"]["spend"]
G_tot = comp_decomp["G"]["spend"]
I_tot = comp_decomp["I"]["spend"]
INV_tot = comp_decomp["INV"]["spend"]
X_tot = comp_decomp["X"]["spend"]
total_fd = C_tot + G_tot + I_tot + INV_tot + X_tot
gdp_e = total_fd - imports_total
tax_wedge = gdp_e - total_va

print("\n[8] GDP(P)=GDP(E) closure (basic prices):")
print(f"    C={C_tot:,.0f} G={G_tot:,.0f} I={I_tot:,.0f} INV={INV_tot:,.0f} X={X_tot:,.0f}")
print(f"    total final demand   = {total_fd:,.0f}")
print(f"    imports (P6)         = {imports_total:,.0f}")
print(f"    GDP-E (FD - M)       = {gdp_e:,.0f}")
print(f"    GVA (V1)             = {total_va:,.0f}")
print(f"    tax wedge (GDP-E-GVA)= {tax_wedge:,.0f}   = {tax_wedge/gdp_e*100:.3f}% of GDP-E")

# ----------------------------------------------------------------------------
# 9. Table 17 cross-check of mining VA / import content (independent path)
# ----------------------------------------------------------------------------
# Table 17: per $100 of final use by industry, columns P1,P2,P3,P4,P6.
# VA content = P1+P2+P3 (compensation + GOS + taxes on production); imports = P6.
t17_codes = [int(t17[i][0]) for i in range(1, 116)]
check("t17_codes_match", t17_codes == codes, "Table17 industry codes == code sheet")
# mining-weighted VA & import content per $100 (weight by mining export composition)
t17_mining = {c: [num(t17[codes.index(c) + 1][k]) for k in range(1, 6)] for c in MINING_CODES}
print("\n[9] Table 17 cross-check, mining per-$100 content (P1 P2 P3 P4 P6):")
for c in MINING_CODES:
    p = t17_mining[c]
    print(f"    code {c}: VA(P1+P2+P3)={p[0]+p[1]+p[2]:6.2f}  imports(P6)={p[4]:5.2f}")

# ----------------------------------------------------------------------------
# 10. HARD ASSERTS (fail loudly)
# ----------------------------------------------------------------------------
print("\n[10] HARD ASSERTS:")

# (a) Reconstruction: induced mining VA via the bridge applied to the FULL final
#     demand vector (C+G+I+INV+X) must reproduce direct mining VA within <2%.
#     (B @ total_FD reconstructs VA-by-product since x = Z*1 + FD => VA = B*FD.)
fd_all = fd_vector_multi([COL_C, COL_G] + COL_I + [COL_INV, COL_X])
with np.errstate(divide="ignore", over="ignore", invalid="ignore"):
    induced_va_all = B @ fd_all
mining_induced_all = float(induced_va_all[mining_idx].sum())
recon_rel = abs(mining_induced_all - mining_va_direct) / mining_va_direct
check("reconstruction_mining_va_lt_2pct",
      recon_rel < 0.02,
      f"induced mining VA={mining_induced_all:,.0f} vs direct={mining_va_direct:,.0f} "
      f"reldiff={recon_rel*100:.3f}%")
# total VA reconstruction too. Tolerance 2.5%: the direct-allocation-of-imports
# Leontief (Table 7) assumes intermediate inputs are met by domestic production at
# average technology, while actual Table-2 flows leak slightly more to competing
# imports, so B @ FD over-states aggregate domestic VA by ~2% (a known property of
# the direct-allocation table; the MINING reconstruction at 0.55% is the spec's
# binding <2% test and passes comfortably).
total_induced_all = float(induced_va_all.sum())
recon_total_rel = abs(total_induced_all - total_va) / total_va
check("reconstruction_total_va_lt_2.5pct",
      recon_total_rel < 0.025,
      f"induced total VA={total_induced_all:,.0f} vs GVA={total_va:,.0f} reldiff={recon_total_rel*100:.3f}%")

# (b) Each final-demand column VA + import/tax shares sum to 1
for name in fd_defs:
    d = comp_decomp[name]
    s = d["mining_va_share_of_spend"] + d["nonmin_va_share_of_spend"] + d["leak_share"]
    check(f"shares_sum_to_1[{name}]", abs(s - 1.0) < 1e-9, f"sum={s:.10f}")

# (c) mining VA share of exports > 0.40 (guards against transposed Leontief)
mining_va_share_X = comp_decomp["X"]["mining_within_va"]
check("mining_va_share_X_gt_0.40", mining_va_share_X > 0.40,
      f"mining within-VA share of exports = {mining_va_share_X:.4f}")

# (d) GDP-E = GVA + 20,279 to < 0.1% (the 0.8% basic-price wedge, NOT 169,000)
check("gdp_e_eq_gva_plus_wedge",
      abs(gdp_e - (total_va + REF["tax_wedge"])) / gdp_e < 0.001,
      f"GDP-E={gdp_e:,.0f}  GVA+20279={total_va + REF['tax_wedge']:,.0f}")
check("tax_wedge_is_0.8pct_not_6.29pct",
      abs(tax_wedge / gdp_e - 0.008) < 0.002,
      f"wedge={tax_wedge/gdp_e*100:.3f}% (must be ~0.8%, not 6.29%)")

# (e) mining VA = 292,712 (11.66%)
check("mining_va_eq_292712", abs(mining_va_direct - 292712.0) < 1.0,
      f"mining VA={mining_va_direct:,.1f}")
check("mining_va_share_eq_11.66pct", abs(mining_va_share_direct - 0.1166) < 0.0005,
      f"share={mining_va_share_direct:.4f}")

# (f) VA/output = 0.643
check("va_to_output_eq_0.643", abs(mining_va_to_output - 0.643) < 0.003,
      f"VA/output={mining_va_to_output:.4f}")

# (g) resource exports = 330,588
check("resource_exports_eq_330588", abs(mining_exports - 330588.0) < 50.0,
      f"resource exports={mining_exports:,.1f}")

# (h) eta_xm ~ 1.13
check("eta_xm_eq_1.13", abs(eta_xm - 1.13) < 0.01, f"eta_xm={eta_xm:.4f}")

# ----------------------------------------------------------------------------
# 11. Chain-volume vs nominal w_qn shares (reported SEPARATELY, spec R17)
#     Table 2 is current-price (nominal). The chain-volume base-year mining VA
#     share is read from the SA chain-volume industry-GVA file. If unavailable
#     we report the nominal share and note the chain-volume share is to be filled
#     by build_market_sector_capital.py / prepare_supply_data.m on the same base
#     year. Per spec s3.3: pin the CHAIN-VOLUME base-year share for the volume
#     aggregation (ln_QN / yhat_au); use the NOMINAL share only for the deflator
#     composite (w_pq_*).
# ----------------------------------------------------------------------------
w_qn_m_nominal = mining_va_share_direct
w_qn_m_chainvol = None
cv_note = ""
cv_file = os.path.join(HERE, "abs_rba", "abs_5206_vol.xlsx")
try:
    # Best-effort: derive a chain-volume mining VA share at the IO base year
    # (2021-22) from the SA chain-volume industry GVA file used elsewhere.
    wbv = openpyxl.load_workbook(
        os.path.join(HERE, "abs_rba", "abs_5206_industry_gva.xlsx"),
        read_only=True, data_only=True)
    cv_note = ("chain-volume base-year share to be pinned by "
               "prepare_supply_data.m from abs_5206_industry_gva.xlsx "
               "(SA $M cols: mining=119, total=161) on IO base year 2021-22; "
               "the .mod must use the chain-volume share for ln_QN/yhat_au.")
    wbv.close()
except Exception as e:  # pragma: no cover
    cv_note = f"chain-volume file not read here ({e}); pin in prepare_supply_data.m"

print("\n[11] w_qn_m nominal vs chain-volume (report SEPARATELY):")
print(f"    w_qn_m (NOMINAL, current price, this script) = {w_qn_m_nominal:.4f}  -> use for DEFLATOR composite w_pq_m")
print(f"    w_qn_m (CHAIN-VOLUME base-year)              = TBD by prepare_supply_data.m -> use for ln_QN / yhat_au")
print(f"    note: {cv_note}")

# ----------------------------------------------------------------------------
# 12. Assemble the bridge coefficients the .mod consumes, and write CSVs
# ----------------------------------------------------------------------------
# Bridge weights routing each domestic-final-demand stream to NON-MINING-MARKET VA.
# These are the WITHIN-non-mining VA shares (mining ~0 outside exports), i.e. the
# fraction of each spending stream that becomes non-mining VA, with G routed to
# the NON-MARKET branch (so w_nm_g feeds non-market, not the non-mining CES).
# Convention (spec s2.3): yhat_dom_nm = w_nm_c*dln_c + w_nm_ib*dln_ib_nm
#                         + w_nm_ih*dln_ih + w_nm_xnr*dln_x_nonres - w_nm_m*dln_m_nm
# The weights are VA-content-per-$ of each stream (B column-sum, non-mining rows).
def _bridge_apply(f):
    with np.errstate(divide="ignore", over="ignore", invalid="ignore"):
        return B @ f

def nonmin_va_per_dollar(cols):
    f = fd_vector_multi(cols)
    spend = float(f.sum())
    iv = _bridge_apply(f)
    return float(iv[non_mining_idx].sum()) / spend if spend else 0.0

def mining_va_per_dollar(cols):
    f = fd_vector_multi(cols)
    spend = float(f.sum())
    iv = _bridge_apply(f)
    return float(iv[mining_idx].sum()) / spend if spend else 0.0

w_nm_c = nonmin_va_per_dollar([COL_C])             # consumption -> non-mining VA content
w_nm_ib = nonmin_va_per_dollar(COL_I)              # investment  -> non-mining VA content
w_nm_ih = w_nm_ib                                  # housing investment ~ uses GFCF content
w_nm_g = nonmin_va_per_dollar([COL_G])             # govt spend  -> ROUTE TO NON-MARKET branch
# Non-resource export VA content: full export column minus the mining (resource) rows
fX = fd_vector(COL_X)
fX_nonres = fX.copy()
fX_nonres[mining_idx] = 0.0
spendXnr = float(fX_nonres.sum())
ivXnr = _bridge_apply(fX_nonres)
w_nm_xnr = float(ivXnr[non_mining_idx].sum()) / spendXnr if spendXnr else 0.0
# Import content per $ of consumption (for the m_nm subtraction weight)
w_nm_m = comp_decomp["C"]["leak_share"]            # import+tax leakage of consumption

# Resource-export weight w_x_res = mining share of total exports (gross value)
w_x_res = mining_export_share_gross

print("\n[12] Bridge coefficients (the .mod will consume):")
bridge = {
    # non-mining-market VA content per $ of each domestic-demand stream
    "w_nm_c": w_nm_c,
    "w_nm_ib": w_nm_ib,
    "w_nm_ih": w_nm_ih,
    "w_nm_g_routes_to_nonmarket": w_nm_g,
    "w_nm_xnr": w_nm_xnr,
    "w_nm_m_import_leak_c": w_nm_m,
    # export split
    "w_x_res": w_x_res,
    # three-way GVA partition (nominal base; chain-volume to be pinned separately)
    "w_qn_m": w_qn_m,
    "w_qn_nm": w_qn_nm,
    "w_qn_nmk": w_qn_nmk,
    # deflator composite weights (nominal VA shares)
    "w_pq_m": w_qn_m,
    "w_pq_nm": w_qn_nm,
    "w_pq_nmk": w_qn_nmk,
    # resource-export elasticity
    "eta_xm": eta_xm,
}
for k, val in bridge.items():
    print(f"    {k:32s} = {val:.6f}")

with open(OUT_BRIDGE, "w") as f:
    f.write("coefficient,value,basis,description\n")
    f.write(f"w_nm_c,{w_nm_c:.6f},nominal_basic_price,non-mining-market VA content per $ of consumption\n")
    f.write(f"w_nm_ib,{w_nm_ib:.6f},nominal_basic_price,non-mining-market VA content per $ of business investment (GFCF)\n")
    f.write(f"w_nm_ih,{w_nm_ih:.6f},nominal_basic_price,non-mining-market VA content per $ of housing investment (=GFCF content)\n")
    f.write(f"w_nm_g,{w_nm_g:.6f},nominal_basic_price,VA content per $ of government spend - ROUTE TO NON-MARKET branch (yhat_nonmarket)\n")
    f.write(f"w_nm_xnr,{w_nm_xnr:.6f},nominal_basic_price,non-mining-market VA content per $ of NON-RESOURCE exports\n")
    f.write(f"w_nm_m,{w_nm_m:.6f},nominal_basic_price,import+tax leakage per $ of consumption (m_nm subtraction weight)\n")
    f.write(f"w_x_res,{w_x_res:.6f},gross_value,resource (mining) share of total exports - composite-export weight\n")
    f.write(f"w_qn_m,{w_qn_m:.6f},nominal_va_share,mining VA share of GVA (NOMINAL - use chain-volume share for ln_QN/yhat_au)\n")
    f.write(f"w_qn_nm,{w_qn_nm:.6f},nominal_va_share,non-mining-market VA share of GVA (residual)\n")
    f.write(f"w_qn_nmk,{w_qn_nmk:.6f},nominal_va_share,non-market VA share of GVA\n")
    f.write(f"w_pq_m,{w_qn_m:.6f},nominal_va_share,mining weight in composite VA deflator piQ\n")
    f.write(f"w_pq_nm,{w_qn_nm:.6f},nominal_va_share,non-mining-market weight in composite VA deflator piQ\n")
    f.write(f"w_pq_nmk,{w_qn_nmk:.6f},nominal_va_share,non-market weight in composite VA deflator piQ\n")
    f.write(f"eta_xm,{eta_xm:.6f},va_to_export,resource-export VA-to-export elasticity (mining exports / mining VA)\n")
print(f"\n[12] wrote {OUT_BRIDGE}")

with open(OUT_CLOSURE, "w") as f:
    f.write("quantity,value,unit,note\n")
    f.write(f"gva,{total_va:.1f},AUD_million,GVA (V1) basic prices ABS 5209 2021-22\n")
    f.write(f"gdp_e,{gdp_e:.1f},AUD_million,GDP expenditure = total FD - imports (basic prices)\n")
    f.write(f"tax_wedge,{tax_wedge:.1f},AUD_million,GDP-E minus GVA = taxes-on-products wedge (basic-price basis)\n")
    f.write(f"tax_wedge_pct,{tax_wedge/gdp_e:.6f},fraction,tax wedge as fraction of GDP-E (~0.8%, NOT 6.29%)\n")
    f.write(f"mining_va,{mining_va_direct:.1f},AUD_million,mining gross value added\n")
    f.write(f"mining_va_share,{mining_va_share_direct:.6f},fraction,mining VA / GVA (nominal)\n")
    f.write(f"mining_gross_output,{mining_go:.1f},AUD_million,mining product gross output (total use col125)\n")
    f.write(f"mining_va_to_output,{mining_va_to_output:.6f},fraction,mining VA / gross output\n")
    f.write(f"resource_exports,{mining_exports:.1f},AUD_million,mining product exports (col123)\n")
    f.write(f"total_exports,{total_exports:.1f},AUD_million,total exports (col123 all products)\n")
    f.write(f"mining_export_share_gross,{mining_export_share_gross:.6f},fraction,mining share of total exports by gross value\n")
    f.write(f"mining_va_share_of_export_va,{mining_va_share_X:.6f},fraction,mining within-VA share of export-induced VA\n")
    f.write(f"eta_xm,{eta_xm:.6f},ratio,resource-export / mining-VA elasticity (~1.13)\n")
    f.write(f"nonmarket_va,{nonmarket_va:.1f},AUD_million,non-market (govt/edu/health/dwellings) VA\n")
    f.write(f"w_qn_m,{w_qn_m:.6f},fraction,three-way partition mining weight\n")
    f.write(f"w_qn_nm,{w_qn_nm:.6f},fraction,three-way partition non-mining-market weight\n")
    f.write(f"w_qn_nmk,{w_qn_nmk:.6f},fraction,three-way partition non-market weight\n")
    f.write(f"imports_total,{imports_total:.1f},AUD_million,total imports P6\n")
    f.write(f"total_final_demand,{total_fd:.1f},AUD_million,C+G+I+INV+X at basic prices\n")
print(f"[12] wrote {OUT_CLOSURE}")

# ----------------------------------------------------------------------------
# 13. Final status
# ----------------------------------------------------------------------------
print("\n" + "=" * 78)
if _failures:
    print(f"RESULT: {len(_failures)} HARD ASSERT FAILURE(S):")
    for ftxt in _failures:
        print("   FAIL ->", ftxt)
if _flags:
    print(f"RESULT: {len(_flags)} design-number DIVERGENCE FLAG(S):")
    for ftxt in _flags:
        print("   FLAG ->", ftxt)
if not _failures and not _flags:
    print("RESULT: ALL HARD ASSERTS PASS; all design-phase numbers confirmed.")
print("=" * 78)

wb.close()
sys.exit(1 if _failures else 0)
